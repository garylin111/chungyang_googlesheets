import streamlit as st
import pygsheets as pg
import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime
from openpyxl.styles import Font
import xlwings as xw


if __name__ == '__main__':
    # 1.數據來源
    st.markdown('# 周排程自動化填寫')
    url = st.text_input('請輸入Google試算表鏈接', key='url')
    font = Font(name="Microsoft JhengHei UI", size=16, color='FF0000', bold=True)
    gc = pg.authorize(service_file=r'C:\Users\asus\Desktop\4_24自動填寫\knife-cost-test.json') # pygsheets只通過本地文件位置尋找金鑰
    week_data = st.file_uploader('周排程文件上傳')
    # 選擇日期
    date_time = st.date_input('選擇填寫日期：', key='date_sel')
    if url:
        sheet = gc.open_by_url(url)
        worksheet = sheet[0]
        data = worksheet.get_all_values()
        df = pd.DataFrame(data[1:], columns=data[0])
        # 選擇保存位置
        address = st.text_input('請輸入你的文件目錄：', key='file_address', value=r'C:\Users\asus\Desktop\4_24自動填寫')
        name = st.text_input('文件名稱：', key='file_name')
        df.rename(columns={'工號姓名 \nTên\nชื่อ': '工號姓名',
                           '班別 \nlớp học \nระดับ ': '班別',
                           '機台編號(例：001) \nSố máy (ví dụ: 001) \nหมายเลขเครื่อง (ตัวอย่าง: 001)': '機台編號',
                           '生產工時(HR)  (例：7.5)\nGiờ sản xuất (HR) (ví dụ: 7.5) \nชั่วโมงการผลิต (HR) (ตัวอย่าง: 7.5)': '生產工時',
                           '生產數量(PCS)   (例：888)\nSố lượng sản xuất (PCS)  (ví dụ: 888) \nปริมาณการผลิต (ชิ้น)  (ตัวอย่าง: 888)': '生產數量',
                           '備註\nLưu ý \nหมายเหตุ ': '備註',
                           '上班日期\nNgày làm việc\nวันที่ทำงาน': '上班日期'}, inplace=True)
        # 将 NaN 替换为有效的列名
        df = df[['工號姓名', '班別', '機台編號', '生產工時', '生產數量', '備註', '上班日期']]
        # 删除列
        if 'NaN_column' in df.columns:
            df.drop(columns=['NaN_column'], inplace=True)

        # 機台編號規整為3位數以上
        for i in range(len(df)):
            if len(str(df['機台編號'].iloc[i])) < 3:
                num_str = str(df['機台編號'][i])
                df['機台編號'][i] = num_str.zfill(3)
        # 將google試算表上工號姓名轉換成只有姓名
        pattern_0 = re.compile(r'[\u4e00-\u9fa5]')
        for i in range(len(df)):
            chinese_characters = re.findall(pattern_0, df['工號姓名'][i])
            connected_chinese = ''.join(chinese_characters)
            df['工號姓名'][i] = connected_chinese
        # 將上班日期轉換爲datatime格式
        for i in range(len(df)):
            date_str = df.loc[i, '上班日期']  # 获取日期字符串
            if date_str and not pd.isna(date_str):  # 检查日期字符串是否为空或者 NaN
                df.loc[i, '上班日期'] = datetime.strptime(date_str, "%Y/%m/%d").date()

        select_data = None
    else:
        st.warning('請輸入試算表鏈接')
    # 2.處理數據
    if st.button('提交', key='submit') and date_time:
        st.write(f'{date_time},請確認這是你需要的時間')
        select_data = df[df['上班日期'] == date_time]
        select_data.replace('', 0, inplace=True)  # 可能出现空字符串，用0替代
        st.dataframe(df[df['上班日期'] == date_time])
        day_time = select_data[select_data['班別'] == '早班人員nhân viên ca sáng(พนักงานกะเช้า)']
        count1 = day_time['工號姓名'].unique()
        night_time = select_data[select_data['班別'] == '晚班人員 Nhân viên ca đêm (พนักงานกะกลางคืน)']
        count2 = night_time['工號姓名'].unique()
        st.write(f'白班人員({len(count1)})：{count1}')
        st.write(f'晚班人員({len(count2)})：{count2}')

        if week_data:
            table_week = load_workbook(week_data)
            worktable = table_week.active
            work_len = worktable.max_row
            work_wide = worktable.max_column

            if not select_data.empty:
                for row in range(9, work_len + 1):
                    for col in range(1, work_wide + 1):
                        cell_value = worktable.cell(row=row, column=col).value
                        if isinstance(cell_value, datetime) and cell_value.date() == date_time:
                            st.write(f"找到匹配的日期：{cell_value.date()}，行 {row}，列 {col}")
                            st.toast('連接機台')
                            for i in range(len(select_data)):
                                key_in = select_data['機台編號'].iloc[i]
                                for j in range(row+1, worktable.max_row):
                                    pattern = re.compile(r'(?<![0-9])[\u4e00-\u9fa5M]*' + re.escape(key_in) + r'[\u4e00-\u9fa5M]*(?![0-9])')
                                    # 判斷 機台
                                    detect_time = 'I' + str(j)
                                    detect_finish = 'E' + str(j)
                                    cell_value = worktable[detect_time].value
                                    # cell_value = worktable.cell(row=j, column=col).value
                                    if cell_value is not None and isinstance(cell_value, (str, int)) and worktable[detect_finish].value != '完工':  # 判斷I欄位内是否有機台編號
                                        if re.search(pattern, str(cell_value)):  # 匹配到google表單機台編號
                                            for k in range(j, j + 14):  # 在同一個機台行，遍歷尋找早晚班實際工時和實際產出
                                                detect_daytime = 'X' + str(k)  # 尋找早晚班實際工時實際產出
                                                # '''人員綁定'''
                                                # 早班
                                                if worktable[detect_daytime].value == '人　　員(早)' and select_data['班別'].iloc[i] == '早班人員nhân viên ca sáng(พนักงานกะเช้า)':
                                                    # '''早班實際人員'''
                                                    if worktable.cell(row=k, column=col).value != select_data['工號姓名'].iloc[i]:
                                                        a_time = select_data['工號姓名'].iloc[i]
                                                        worktable.cell(row=k, column=col, value=a_time).font = font
                                                    else:
                                                        a_time = select_data['工號姓名'].iloc[i]
                                                        worktable.cell(row=k, column=col, value=a_time)
                                                # 晚班
                                                elif worktable[detect_daytime].value == '人　　員(晚)' and select_data['班別'].iloc[i] == '晚班人員 Nhân viên ca đêm (พนักงานกะกลางคืน)':
                                                    # '''晚班實際人員'''
                                                    if worktable.cell(row=k, column=col).value != select_data['工號姓名'].iloc[i]:
                                                        a_time = select_data['工號姓名'].iloc[i]
                                                        worktable.cell(row=k, column=col, value=a_time).font = font
                                                    else:
                                                        a_time = select_data['工號姓名'].iloc[i]
                                                        worktable.cell(row=k, column=col, value=a_time)
                                                elif worktable[detect_daytime].value == '實際產出(早)' and select_data['班別'].iloc[i] == '早班人員nhân viên ca sáng(พนักงานกะเช้า)':
                                                    # '''早班產出'''
                                                    if worktable.cell(row=k, column=col).value is not None:
                                                        a_time = float(select_data['生產數量'].iloc[i])
                                                        act_time = a_time + float(worktable.cell(row=k, column=col).value)
                                                        worktable.cell(row=k, column=col, value=act_time)
                                                    else:
                                                        a_time = float(select_data['生產數量'].iloc[i])
                                                        worktable.cell(row=k, column=col, value=a_time)
                                                elif worktable[detect_daytime].value == '實際產出(晚)' and select_data['班別'].iloc[i] == '晚班人員 Nhân viên ca đêm (พนักงานกะกลางคืน)':
                                                    # '''晚班產出'''
                                                    if worktable.cell(row=k, column=col).value is not None:
                                                        a_time = float(select_data['生產數量'].iloc[i])
                                                        act_time = a_time + float(worktable.cell(row=k, column=col).value)
                                                        worktable.cell(row=k, column=col, value=act_time)
                                                    else:
                                                        a_time = float(select_data['生產數量'].iloc[i])
                                                        worktable.cell(row=k, column=col, value=a_time)
                                                elif worktable[detect_daytime].value == '實際工時(早)' and select_data['班別'].iloc[i] == '早班人員nhân viên ca sáng(พนักงานกะเช้า)':
                                                    # '''早班工時'''
                                                    if worktable.cell(row=k, column=col).value is not None:
                                                        a_time = float(select_data['生產工時'].iloc[i])
                                                        act_time = a_time + float(worktable.cell(row=k, column=col).value)
                                                        worktable.cell(row=k, column=col, value=act_time)
                                                    else:
                                                        a_time = float(select_data['生產工時'].iloc[i])
                                                        worktable.cell(row=k, column=col, value=a_time)
                                                elif worktable[detect_daytime].value == '實際工時(晚)' and select_data['班別'].iloc[i] == '晚班人員 Nhân viên ca đêm (พนักงานกะกลางคืน)':
                                                    # '''晚班工時'''
                                                    if worktable.cell(row=k, column=col).value is not None:
                                                        a_time = float(select_data['生產工時'].iloc[i])
                                                        act_time = a_time + float(worktable.cell(row=k, column=col).value)
                                                        worktable.cell(row=k, column=col, value=act_time)
                                                    else:
                                                        a_time = float(select_data['生產工時'].iloc[i])
                                                        worktable.cell(row=k, column=col, value=a_time)
                                                elif worktable[detect_daytime].value == '備註說明(早)' and select_data['班別'].iloc[i] == '早班人員nhân viên ca sáng(พนักงานกะเช้า)':
                                                    a_time = select_data['備註'].iloc[i]
                                                    worktable.cell(row=k, column=col, value=a_time)
                                                elif worktable[detect_daytime].value == '備註說明(晚)' and select_data['班別'].iloc[i] == '晚班人員 Nhân viên ca đêm (พนักงานกะกลางคืน)':
                                                    a_time = select_data['備註'].iloc[i]
                                                    worktable.cell(row=k, column=col, value=a_time)

        # '''加上一個循環，通過前面得到的總工時，通過遍歷計算總工時/實際機台數 = 實際機台運行時間'''
            st.toast('遍歷計算總工時/實際機台數 = 實際機台運行時間')
            for row in range(9, work_len + 1):
                for col in range(1, work_wide + 1):
                    cell_value = worktable.cell(row=row, column=col).value
                    if isinstance(cell_value, datetime) and cell_value.date() == date_time:
                        # '''機台連接'''
                        for j in range(row, worktable.max_row):  # 從第10行開始
                            # '''判斷 機台(特殊情況：005，12345，004直接跳過這個機台的實際時間平均)'''
                            pattern_1 = re.compile(r'005|12345|004')
                            detect_time = 'I' + str(j)
                            cell_value = worktable[detect_time].value
                            if cell_value is not None:
                                cell_value_str = str(cell_value)
                                if re.search(pattern_1, cell_value_str):
                                    j = j + 1
                                # cell_value = worktable[detect_time].value
                            # cell_value = worktable.cell(row=j, column=col).value
                                elif isinstance(cell_value, str):  # 判斷I欄位内是否有機台編號
                                    detect_mach = 'M' + str(j)  # 機台數量
                                    if isinstance(worktable[detect_mach].value, (int, float)):
                                        mach_run_num = float(worktable[detect_mach].value)
                                        for k in range(j, j + 15):  # 在同一個機台行，遍歷尋找早晚班實際工時
                                            detect_daytime = 'X' + str(k)  # 尋找早晚班實際工時
                                            if worktable[detect_daytime].value == '實際工時(早)':
                                                # '''早班工時'''
                                                if worktable.cell(row=k, column=col).value is not None:
                                                    if worktable[detect_mach].value is not None:
                                                        time_work = float(worktable.cell(row=k, column=col).value)
                                                        act_time_work = round(time_work / mach_run_num, 2)  # 保留一位小數
                                                        worktable.cell(row=k, column=col, value=act_time_work)
                                            elif worktable[detect_daytime].value == '實際工時(晚)':
                                                # ''''晚班工時'''
                                                if worktable.cell(row=k, column=col).value is not None:
                                                    if worktable[detect_mach].value is not None:
                                                        time_work = float(worktable.cell(row=k, column=col).value)
                                                        act_time_work = round(time_work / mach_run_num, 2)  # 保留一位小數
                                                        worktable.cell(row=k, column=col, value=act_time_work)
    # 4.寫入數據
            file_name = '\\' + name + '.xlsx'
            if address and file_name and name:
                table_week.save(address + file_name)

                excel_app = xw.App(visible=False)
                excel_book = excel_app.books.open(address + file_name)
                excel_book.save()
                excel_book.close()
                excel_app.quit()

                st.success('生成成功！' + address + file_name)
                st.balloons()